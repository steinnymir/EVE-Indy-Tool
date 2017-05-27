# -*- coding: utf-8 -*-
"""
Created on Sat May 20 17:02:12 2017

@author: Stymir
"""
from openpyxl import load_workbook
from library import gfs
import csv
import os
import yaml
import pickle


def main():
    """ this refreshes import of data from yaml sde to pickle format"""
    sde = SDE()
    sde.import_quick()
    # dbList = ['blueprints', 'typeIDs']
    # # import yaml / csv and dump pickle
    # sde.export_multiple_pickle(dbList)
    name = sde.typeIDs[34]['name']['en']

    ID = sde.get_ID_from_name('condor')

    print(sde.get_parent_BP(ID))

class SDE(object):
    """ Contains a series of dictionaries containing necessary sde data"""

    def __init__(self):

        DB_LOCATION = '../database/'
        DB_LOCATION =  'D:/Documents/py_code/EVEIndyTool/database/'  # comment when at home

        self.DB_LOCATION_CSV = 'D:/Documents/py_code/EVE_database/' + 'csv/'
        self.DB_LOCATION_YAML = 'D:/Documents/py_code/EVE_database/' + 'sde/fsd/'
        self.DB_LOCATION_PICKLE = DB_LOCATION
        self.QUICK_IMPORT_LIST = ['typeIDs', 'blueprints']

        self.db_list = []

        self.blueprints = None
        self.invCategories = None
        self.invGroups = None
        self.invMetaGroups = None
        self.invMetaTypes = None
        self.invTypes = None
        self.typeIDs = None


        self.db_coreIndy = ['invTypes',
                            'invCategories',
                            'invGroups',
                            'invMetaTypes',
                            'invMetaGroups',
                            'blueprints']

        self.db_coreIndy_old = ['invTypes',
                                'invCategories',
                                'invGroups',
                                'invMetaTypes',
                                'invMetaGroups',
                                'industryActivity',
                                'industryActivityMaterials',
                                'industryActivityProbabilities']

        self.activityTypeID = {0: 'None',
                               1: 'Manufacturing',
                               2: 'Researching Technology',
                               3: 'Researching Time Efficiency',
                               4: 'Researching Material Efficiency',
                               5: 'Copying',
                               6: 'Duplicating',
                               7: 'Reverse Engineering',
                               8: 'Invention'}

    def import_data(self, dbName):
        """ imports data from a database file (csv or yaml)"""
        timer = gfs.Timer()
        timer.tic()
        if dbName in ['blueprints', 'typeIDs', 'groupIDs', 'graphicIDs']:
            print('importing: ' + dbName + '.yaml')
            data = self.get_data_yaml(dbName)
        else:
            print('importing: ' + dbName + '.csv')
            data = self.get_data_csv(dbName)
        setattr(self, dbName, data)
        dt = timer.toc(out='return')
        print('Imported {0} in {1:.3f} ms'.format(dbName, dt))

    def export_pickle(self, dbName):
        """ create a pickle for each database loaded"""
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        try:
            timer = gfs.Timer()
            timer.tic()
            data = getattr(self, dbName)
            with open(filePath, 'wb+') as f:
                pickle.dump(data, f)
            dt = timer.toc(out='return')
            print('Exported {0} in {1:.3f} ms'.format(dbName, dt))
        except AttributeError:
            print('no data found with name ' + dbName)

    def export_multiple_pickle(self, dbList):
        """ import from sde and export to pickle the list of databases given """
        for db_name in dbList:
            print('importing: ' + db_name)
            self.import_data(db_name)
            print('exporting: ' + db_name)
            self.export_pickle(db_name)

    def import_pickle(self, dbName):
        """ imports data from pickle dumped file"""
        # try:
        timer = gfs.Timer()
        timer.tic()
        print('Importing: ' + dbName)
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        with open(filePath, 'rb') as f:
            data = pickle.load(f)
            setattr(self, dbName, data)
        dt = timer.toc(out='return')
        print('Imported {0} in {1:.3f} ms'.format(dbName, dt))

    def import_quick(self):
        """ import databases relevant for indy calculator"""
        for dbName in self.QUICK_IMPORT_LIST:
            self.import_pickle(dbName)

    def get_filepath(self, db_name, ext):
        """
        :return : returns str of database file path.
        :param db_name: name of database file to load
        :type ext: str representing db_name file format

        """
        if ext == 'yaml':
            return self.DB_LOCATION_CSV + db_name + '.' + ext
        elif ext == 'csv':
            return self.DB_LOCATION_YAML + db_name + '.' + ext
        else:
            return 'Incorrect file extension.'

    def get_data_yaml(self, db_name):
        """ imports a YAML database to same name attribute"""
        with  open(self.DB_LOCATION_YAML + db_name + '.yaml', 'r', encoding="utf8") as f:
            data = yaml.load(f)
        return data

    def get_data_csv(self, db_name):
        """ imports a CSV database to same name attribute """
        headline = True
        data = {}
        with open(self.DB_LOCATION_CSV + db_name + '.csv', 'r', encoding="utf8") as f:
            reader = csv.reader(f)

            for line in reader:  # get headers for dict key names
                if headline:
                    headers = line
                    headline = False
                else:
                    key = int(line[0])
                    pos = 0

                    try:  # try appending new values to repeated 1st column items
                        # otherwise just add new entry
                        _ = data[key]
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = gfs.getNum_or_Str(word)
                            try:
                                data[key][label].append(val)
                            except AttributeError:
                                tmp = data[key][label]
                                data[key][label] = [tmp]
                                data[key][label].append(val)

                    except KeyError:
                        data[key] = {}
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = gfs.getNum_or_Str(word)

                            data[key][label] = val
        return (data)

    def get_ID_from_name(self, name):
        """ returns the itemID from a given name

        :return: int
        """
        itemName = ''
        for key in self.typeIDs:
            try:
                itemName = str(self.typeIDs[key]['name']['en']).lower()
                if itemName == name.lower():
                    return key
            except KeyError:
                pass


    def get_parent_BP(self, itemID):  # todo: Fix me
        """ returns the blueprintID that produces the item with given itemID
        :return: int
        """
        parent_tipeID = 0
        for key in self.blueprints:
            try:
                product_list = self.blueprints[key]['activities']['manufacturing']['products']
                for item in product_list:  # products is a list. so we need to break the dcit style and add list parser
                    parent_tipeID = item['typeID']
                    if parent_tipeID == itemID:
                        return key
            except KeyError:
                pass



            # --------------- Old and deprectated methods ------------------


    def open_file_xlsx(self, file):
        ''' '''
        wb = load_workbook(file)
        ws = wb.active
        self.name = ws['A1'].value
        # get column headers
        for col in ws.iter_cols(max_row=1):
            for cell in col:
                self.headers.append(cell.value)

        # write to data
        row_i = 0

        for row in ws.iter_rows(min_row=2):

            row_i += 1
            key = ws.cell(row=row_i, column=1).value
            col_i = 0
            self.data[key] = {}
            for header in self.headers:
                col_i += 1
                self.data[key][header] = ws.cell(row=row_i, column=col_i).value


if __name__ == '__main__':
    main()
