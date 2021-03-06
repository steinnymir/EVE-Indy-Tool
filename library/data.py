# -*- coding: utf-8 -*-
"""
Created on Sat May 20 17:02:12 2017

@author: Steinn Ymir
"""

import csv
import json
import os
import sys
import pickle
from urllib import request
from xml.dom import minidom
import configparser
import yaml
from openpyxl import load_workbook
import time
from library import gfs


def main():
    """ this refreshes import of data from yaml sde to pickle format"""
    # sde = SDE()
    # sde.import_and_export()
    market = Market()
    market.update_marketData()

class SDE(object):
    """ Contains a series of dictionaries containing necessary sde data"""

    def __init__(self):
        """ initialize attributes where to store all sde data"""
        parser = configparser.ConfigParser()

        parser.read('D:/Documents/py_code/EVEIndyTool/settings.ini')
        self.DB_LOCATION_CSV = parser.get('locations', 'DB_LOCATION_CSV')
        self.DB_LOCATION_PRIMARY = parser.get('locations', 'DB_LOCATION_PRIMARY')
        self.DB_LOCATION_SECONDARY = parser.get('locations', 'DB_LOCATION_SECONDARY', )
        self.DB_LOCATION_PICKLE = 'D:/Documents/py_code/EVEIndyTool/database/'

        #self.QUICK_IMPORT_LIST = ('typeIDs', 'blueprints', 'categoryIDs', 'groupIDs')
        self.PRIMARY_IMPORT_LIST = ('typeIDs', 'blueprints', 'categoryIDs', 'groupIDs', 'iconIDs')
        self.SECONDARY_IMPORT_LIST = (
            'invMarketGroups', 'invMetaGroups', 'invMetaTypes', 'invNames', 'invTypeMaterials', 'ramActivities',
            'ramAssemblyLineStations', 'ramAssemblyLineTypeDetailPerCategory', 'ramAssemblyLineTypeDetailPerGroup',
            'ramAssemblyLineTypes', 'ramInstallationTypeContents')

        self.db_list = []
        # primary
        try:
            self.typeIDs = self.fetch_data_from_pickle('typeIDs')
        except:
            self.typeIDs = None
        try:
            self.blueprints = self.fetch_data_from_pickle('blueprints')
        except:
            self.blueprints = None
        try:
            self.categoryIDs = self.fetch_data_from_pickle('categoryIDs')
        except:
            self.categoryIDs = None

        try:
            self.groupIDs = self.fetch_data_from_pickle('groupIDs')
        except:
            self.groupIDs = None
        try:
            self.invMarketGroups = self.fetch_data_from_pickle('invMarketGroups')
        except:
            self.invMarketGroups = None
        try:
            self.invMetaGroups = self.fetch_data_from_pickle('invMetaGroups')
        except:
            self.invMetaGroups = None
        try:
            self.invMetaTypes = self.fetch_data_from_pickle('invMetaTypes')
        except:
            self.invMetaTypes = None
        try:
            self.invNames = self.fetch_data_from_pickle('invNames')
        except:
            self.invNames = None
        try:
            self.invTypeMaterials = self.fetch_data_from_pickle('invTypeMaterials')
        except:
            self.invTypeMaterials = None
        try:
            self.ramActivities = self.fetch_data_from_pickle('ramActivities')
        except:
            self.ramActivities = None
        try:
            self.ramAssemblyLineStations = self.fetch_data_from_pickle('ramAssemblyLineStations')
        except:
            self.ramAssemblyLineStations = None
        try:
            self.ramAssemblyLineTypeDetailPerCategory = self.fetch_data_from_pickle('ramAssemblyLineTypeDetailPerCategory')
        except:
            self.ramAssemblyLineTypeDetailPerCategory = None
        try:
            self.ramAssemblyLineTypeDetailPerGroup = self.fetch_data_from_pickle('ramAssemblyLineTypeDetailPerGroup')
        except:
            self.ramAssemblyLineTypeDetailPerGroup = None
        try:
            self.ramAssemblyLineTypes = self.fetch_data_from_pickle('ramAssemblyLineTypes')
        except:
            self.ramAssemblyLineTypes = None

        try:
            self.ramInstallationTypeContents = self.fetch_data_from_pickle('ramInstallationTypeContents')
        except:
            self.ramInstallationTypeContents = None


    def import_and_export(self):
        """ Import all database data from yaml and dump to pickle
        """
        timer = gfs.Timer()
        timer.tic()
        print('\nimporting all data:\n')
        self.import_all_yaml()
        timer.toc()
        print('\ndumping all data:\n')
        self.dump_all_as_pickle()
        timer.toc()
        print('\ndump successful\n')

    def import_all_yaml(self):

        for dbName in self.PRIMARY_IMPORT_LIST:
            filepath = self.DB_LOCATION_PRIMARY + dbName + '.yaml'
            self.import_data(filepath)

        for dbName in self.SECONDARY_IMPORT_LIST:
            filepath = self.DB_LOCATION_SECONDARY + dbName + '.yaml'
            self.import_data(filepath)

    def dump_all_as_pickle(self):
        """ dumps all loaded data to pickle databases"""
        for dbName in self.PRIMARY_IMPORT_LIST:
            if dbName is not None:
                self.export_pickle(dbName)
        for dbName in self.SECONDARY_IMPORT_LIST:
            if dbName is not None:
                self.export_pickle(dbName)

    def reload_all(self):
        """ loads all available pickle format databases"""
        dbList = self.PRIMARY_IMPORT_LIST + self.SECONDARY_IMPORT_LIST
        for dbName in dbList:
            self.import_pickle(dbName)

    def import_data(self, dbFilepath):
        """ imports data from a database file (csv or yaml)"""
        timer = gfs.Timer()
        timer.tic()
        dbName = os.path.basename(dbFilepath)
        dbExt = dbName.split('.')[-1]
        data = None
        if dbExt == 'yaml':
            print('importing: ' + dbName)
            data = self.get_data_yaml(dbFilepath)
        elif dbExt == 'csv':
            print('importing: ' + dbName)
            data = self.get_data_csv(dbName)
        else:
            print('Import failed: unrecognised format for file {}'.format(dbName))
        if data is not None:
            setattr(self, dbName.split('.')[0], data)
            dt = timer.toc(out='return')
            print('Imported {0} in {1:.3f} ms'.format(dbName, dt))

    def export_pickle(self, dbName):
        """ create a pickle for each database loaded"""
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        print('Exporting {0} to Pickle database'.format(filePath))
        try:
            timer = gfs.Timer()
            timer.tic()
            data = getattr(self, dbName)
            with open(filePath, 'wb+') as f:
                pickle.dump(data, f)
            dt = timer.toc(out='return')
            print('Exported {0} in {1:.3f} ms'.format(dbName, dt))
        except AttributeError:
            print('no data found with name ' + dbName)

    def import_pickle(self, dbName):
        """ imports data from pickle dumped file"""
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        timer = gfs.Timer()
        timer.tic()
        try:
            print('Importing: ' + dbName)
            with open(filePath, 'rb') as f:
                data = pickle.load(f)
                setattr(self, dbName, data)
            dt = timer.toc(out='return')
            print('Imported {0} in {1:.3f} ms'.format(dbName, dt))
        except FileNotFoundError:
            print('Error 404: file not found - ' + filePath)

    def fetch_data_from_pickle(self, dbName):
        """ imports data from pickle dumped file"""
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        timer = gfs.Timer()
        timer.tic()
        try:
            print('Importing: ' + dbName)
            with open(filePath, 'rb') as f:
                data = pickle.load(f)
            dt = timer.toc(out='return')
            print('Imported {0} in {1:.3f} ms'.format(dbName, dt))
            return data
        except FileNotFoundError:
            print('Error 404: file not found - ' + filePath)


    def get_data_yaml(self, dbFilepath):
        """ imports a YAML database to same name attribute"""
        with  open(dbFilepath, 'r', encoding="utf8") as f:
            data = yaml.load(f)
        return data

    def get_data_csv(self, db_name):
        """ imports a CSV database to same name attribute """
        headline = True
        data = {}
        with open(self.DB_LOCATION_CSV + db_name + '.csv', 'r', encoding="utf8") as f:
            reader = csv.reader(f)

            for line in reader:  # get headers for dict key names
                if headline:
                    headers = line
                    headline = False
                else:
                    key = int(line[0])
                    pos = 0

                    try:  # try appending new values to repeated 1st column items
                        # otherwise just add new entry
                        _ = data[key]
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = gfs.getNum_or_Str(word)
                            try:
                                data[key][label].append(val)
                            except AttributeError:
                                tmp = data[key][label]
                                data[key][label] = [tmp]
                                data[key][label].append(val)

                    except KeyError:
                        data[key] = {}
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = gfs.getNum_or_Str(word)

                            data[key][label] = val
        return data

    # ++++++++ Functionalities +++++++++++

    def get_ID_from_name(self, name):
        """ returns the itemID from a given name
        :return: int
        """
        itemName = ''
        for key in self.typeIDs:
            try:
                itemName = str(self.typeIDs[key]['name']['en']).lower()
                if itemName == name.lower():
                    return key
            except KeyError:
                return None

    def get_parent_blueprintID(self, itemID):  # todo: Fix me
        """ returns the blueprintID that produces the item with given itemID, if none, return None
        """
        item_ID = int(itemID)
        for key in self.blueprints:
            try:
                product_list = self.blueprints[key]['activities']['manufacturing']['products']
                for item in product_list:  # products is a list. so we need to break the dict style and add list parser
                    parent_typeID = int(item['typeID'])
                    if parent_typeID == item_ID:
                        return key
            except KeyError:
                pass
        return None

    def get_parent_invention_blueprintID(self, blueprintID):
        """ return the blueprint which invented gives the input blueprintID, if none, return None"""
        for key in self.blueprints:
            try:
                product_list = self.blueprints[key]['activities']['invention']['products']
                for item in product_list:  # products is a list. so we need to break the dict style and add list parser
                    parent_typeID = item['typeID']
                    if parent_typeID == blueprintID:
                        return key
            except KeyError:
                pass
        return None


                # --------------- Old and deprectated methods ------------------

    def open_file_xlsx(self, file):
        """ """
        wb = load_workbook(file)
        ws = wb.active
        self.name = ws['A1'].value
        # get column headers
        for col in ws.iter_cols(max_row=1):
            for cell in col:
                self.headers.append(cell.value)

        # write to data
        row_i = 0

        for row in ws.iter_rows(min_row=2):

            row_i += 1
            key = ws.cell(row=row_i, column=1).value
            col_i = 0
            self.data[key] = {}
            for header in self.headers:
                col_i += 1
                self.data[key][header] = ws.cell(row=row_i, column=col_i).value


class API(object):
    """ """
    number_of_characters = 12

    def __init__(self):
        """ """
        self.AccountBalance = {}
        self.AssetList = {}
        self.Blueprints = {}
        self.IndustryJobs = {}
        self.MarketOrders = {}
        self.skills = {}
        self.char_api = {}
        self.corp_api = {}
        self.current_key = 1
        self.apikey = []
        self.api_URLs = {'AssetList': "https://api.eveonline.com/corp/AssetList.xml.aspx?flat=1&",
                         'Blueprints': "https://api.eveonline.com/corp/Blueprints.xml.aspx?",
                         'MarketOrders': "https://api.eveonline.com/corp/MarketOrders.xml.aspx?",
                         'IndustryJobs': "https://api.eveonline.com/corp/IndustryJobs.xml.aspx?",
                         'AccountBalance': "https://api.eveonline.com/corp/AccountBalance.xml.aspx?",
                         'Skills': 'https://api.eveonline.com//char/Skills.xml.aspx?'
                         }
        self.fetch_api()
        self.fetch_api(corp=False)

    def fetch_api(self, corp=True):
        """ get api values from keys.ini. File must be in main program directory"""
        char_number = 12
        parser = configparser.ConfigParser()
        parser.read('D:\Documents\py_code\EVEIndyTool\keys.ini')
        if corp:
            api_type = 'api_corp'
        else:
            api_type = 'api_char'
        for i in range(char_number):
            try:
                name = parser.get(api_type, 'char{}_name'.format(i+1))
                characterID = parser.get(api_type, 'char{}_characterID'.format(i+1))
                KeyID = parser.get(api_type, 'char{}_KeyID'.format(i+1))
                vCode = parser.get(api_type, 'char{}_vCode'.format(i+1))

                if corp:
                    self.corp_api[name] = {}
                    self.corp_api[name]['KeyID'] = KeyID
                    self.corp_api[name]['vCode'] = vCode
                    self.corp_api[name]['characterID'] = characterID
                else:
                    self.char_api[name] = {}
                    self.char_api[name]['KeyID'] = KeyID
                    self.char_api[name]['vCode'] = vCode
                    self.char_api[name]['characterID'] = characterID
            except configparser.NoOptionError:
                pass

    def iterate_keys(self):
        """ iterate between keys to use"""
        parser = configparser.ConfigParser()
        parser.read('../keys.ini')
        if self.current_key == self.number_of_characters:
            self.current_key = 1
        else:
            self.current_key += 1
        name = parser.get('api', 'char{}_name'.format(self.current_key))
        KeyID = parser.get('api', 'char{}_KeyID'.format(self.current_key))
        vCode = parser.get('api', 'char{}_vCode'.format(self.current_key))
        self.apikey = (name, KeyID, vCode)

    def get_key_by_character_name(self, character):
        """ iterate between keys to use"""
        parser = configparser.ConfigParser()
        parser.read('../keys.ini')
        for i in range(self.number_of_characters):
            name = parser.get('api', 'char{}_name'.format(self.current_key))
            KeyID = parser.get('api', 'char{}_KeyID'.format(self.current_key))
            vCode = parser.get('api', 'char{}_vCode'.format(self.current_key))
            if name == character:
                return (KeyID, vCode)
        print('no character named ' + character)

    def update_AssetList(self):  # todo: add cachedUntill check
        # todo: check keys for optimized data import
        self.AssetList = self.fetch_eveapi_data('AssetList')

    def update_Blueprints(self):
        self.Blueprints = self.fetch_eveapi_data('Blueprints')

    def update_MarketOrders(self):
        self.MarketOrders = self.fetch_eveapi_data('MarketOrders')

    def update_IndustryJobs(self):
        self.IndustryJobs = self.fetch_eveapi_data('IndustryJobs')

    def update_AccountBalance(self):
        self.AccountBalance = self.fetch_eveapi_data('AccountBalance')

    def update_Skills(self):  # todo: get character api from pedro
        """ get all api keys, and get each char's skill list into a dict"""
        for character in self.char_api:
            url = str(self.api_URLs['Skills'] +
                      'characterID=' + str(self.char_api[character]['characterID']) +
                      '&keyID=' + str(self.char_api[character]['KeyID']) +
                      '&vCode=' + str(self.char_api[character]['vCode']))
            self.skills[character] = self.fetch_eveapi_data('Skills', url=url)



    def update_All(self):
        self.update_AccountBalance()
        self.update_AssetList()
        self.update_Blueprints()
        self.update_IndustryJobs()
        self.update_MarketOrders()
        self.iterate_keys()

    def get_api_url(self, api_type):  # todo: change api choice to api_corp / api_char dictionaries
        """ """
        url = str(self.api_URLs[api_type] +
                  'keyID=' + str(self.apikey[0]) +
                  '&vCode=' + str(self.apikey[1]))
        return url

    def fetch_eveapi_data(self, api_type, url=None, dict_key=None):  # todo: check structure

        """ get data from api.
        structure: {dict_key = {dict_data:val,dict_data,val}}"""
        if url is None:
            url = self.get_api_url(api_type)

        print('Requesting {} data from API system'.format(api_type))
        print(url)
        xmldoc = minidom.parse(request.urlopen(url))  # parse the url, fetch data from xml into xmldoc
        try:
            currentTime = xmldoc.getElementsByTagName('currentTime')[0].firstChild.nodeValue
            cachedUntil = xmldoc.getElementsByTagName('cachedUntil')[0].firstChild.nodeValue

            data_dict = {'times': {'cachedUntil': cachedUntil,  # initialize output dictionary appending time medatada
                                   'currentTime': currentTime}}
        except:
            pass

        header_line = xmldoc.getElementsByTagName('rowset')  # get xml table legend
        api_name = header_line[0].attributes['name'].value

        if dict_key is None:  # unless specified, assign value in key as title for sub-dictionaries in data output
            dict_key = header_line[0].attributes['key'].value
        column_headers = header_line[0].attributes['columns'].value.split(',')
        data_body = xmldoc.getElementsByTagName('row')

        for line in data_body:
            data_dict[line.attributes[dict_key].value] = {}
            for col in column_headers:
                data_dict[line.attributes[dict_key].value][col] = line.attributes[col].value

        print('successfully imported {} data'.format(api_type))
        return data_dict


class ESI(object):  # todo: find out how to make authorized requests
    """ class managing esi data download"""

    def __init__(self):
        """ """
        self.url_serverStatus = 'https://esi.tech.ccp.is/latest/status/?datasource=tranquility&user_agent=eveindytool'
        self.DB_LOCATION = 'D:/Documents/py_code/EVEIndyTool/database/'
        self.root = 'https://esi.tech.ccp.is/latest/'
        self.user_agent = 'eveindytool_Pax_Correl'
        DB_LOCATION = 'D:/Documents/py_code/EVEIndyTool/database/'  # comment when at home

    def fetch_esi_data(self, url=None):  # todo: cached time check and error management

        if url is None:
            url = self.url_serverStatus
        j = request.urlopen(url)
        data = json.load(j)
        return data


class Market(ESI):
    """ the eve market"""

    def __init__(self):
        super().__init__()
        self.data = {}

        try:
            self.load_marketData()
        except EOFError:
            pass
        except FileNotFoundError:
            print('no market data, please update with update_marketData')

    def get_min_sellprice(self, item_id, stationID=60003760):  # todo: expand functionalities as min volume etc
        """ returns the minimum sell price for given item_id"""
        minprice = 1000000000000
        for key in self.data[item_id]:
            if not self.data[item_id][key]['is_buy_order']:
                price = self.data[item_id][key]['price']
                loc = self.data[item_id][key]['location_id']
                if loc == stationID:  # if it is in jita 4-4
                    minprice = min([minprice, price])
        return minprice

    def load_marketData(self):
        filePath = self.DB_LOCATION + 'marketdata.p'
        print('Importing: marketdata')
        timer = gfs.Timer()
        timer.tic()
        with open(filePath, 'rb') as f:
            self.data = pickle.load(f)
        dt = timer.toc(out='return')
        print('Imported marketdata in {0:3f} ms'.format(dt))


    def update_marketData(self):
        """ update from web and dump to pickle database"""
        data = self.get_market_data()
        print(type(data))
        print(data)
        self.data = data
        filePath = self.DB_LOCATION + 'marketdata.p'

        print('dumping data to ' + filePath)
        with open(filePath, 'wb+') as f:
            pickle.dump(data, f)

    def import_price_history(self, type_id, location=10000002):
        """ get price statistics for given item in given location, default is The Forge"""

        request_url = (self.root + 'markets' + str(location) +
                       'history/?datasource=tranquility&type_id=' +
                       type_id + '&user_agent=' + self.user_agent)
        return self.fetch_esi_data(request_url)

    def get_market_data(self, itemID=None, location=10000002, ordertype='all', maxpages=0):
        """ return dict of full market data,
        location: default location The Forge
        ordertype: 'buy' 'sell' or 'all' - only works when typeID is given
        maxpages: number of pages to download, 0 means all

        :returns data_dict, dictionary in format: {itemID : {sell: {0: data, 1:data},
                                                             buy{0: data, 1:data}}}"""

        pagenum = 0
        got_empty_page = False
        data_list = []
        data_dict = {}
        print('requesting Market data from locationID: ' + str(location) + ' through ESI')
        while not got_empty_page:
            pagenum += 1
            request_url = (self.root + 'markets/' + str(location) +
                           '/orders/?datasource=tranquility' +
                           '&order_type=' + ordertype +
                           '&page=' + str(pagenum))
            if itemID is not None:
                request_url += '&type_id=' + str(itemID)
            request_url += '&user_agent=' + self.user_agent
            print('page ' + str(pagenum))
            print(request_url)
            result = self.fetch_esi_data(request_url)
            # when result contains less than the 10000 limit per page, stop iterating pages
            if len(result) != 10000 or pagenum == maxpages or itemID is not None:
                got_empty_page = True
            for item in result:
                data_list.append(item)
        print(str(len(data_list)) + ' market entries imported.')

        # sort results in a better dictionary
        for item in data_list:
            try:
                num = len(data_dict[item['type_id']].keys())
                data_dict[item['type_id']][num + 1] = item
            except KeyError:
                data_dict[item['type_id']] = {1: item}
        print(data_dict)
        #
        #
        # for item in data_list:
        #     if item['is_buy_order']:
        #         try:
        #             buynum = len(data_dict[item['type_id']]['buy'].keys())
        #             data_dict[item['type_id']]['buy'][str(buynum + 1)] = item
        #         except KeyError:
        #             data_dict[item['type_id']] = {'buy': {'0': item}}
        #     else:
        #         try:
        #             sellnum = len(data_dict[item['type_id']]['sell'].keys())
        #             data_dict[item['type_id']]['sell'][str(sellnum + 1)] = item
        #         except KeyError:
        #             data_dict[item['type_id']] = {'sell': {'0': item}}
        return data_dict


if __name__ == '__main__':
    os.chdir('../')
    main()
