# -*- coding: utf-8 -*-
"""
Created on Sat May 20 17:02:12 2017

@author: Stymir
"""
from openpyxl import load_workbook
# from . import gfs
from library import gfs
import csv
import os
import yaml
import pickle

try:
    from yaml import CLoader as Loader, CDumper as Dumper
except ImportError:
    from yaml import Loader, Dumper


def main():
    timer = gfs.Timer()
    timer.tic()

    # test_bp = Blueprint(28675)
    # test_bp.printName()
    # test_bp.fetch_bp_data()
    # print(test_bp.get_manufacturing_materials(20))

    # sde = SDE()
    # sde.import_data('blueprints')
    # sde.export_pickle('blueprints')
    #
    # timer.toc()







class EVEItem(object):
    """ an object in new eden """

    def __init__(self, itemID, SDE):
        """ initialzie"""
        self.itemID = itemID
        self.itemName = '' # todo: make the id to name interpreter

        self.initialize_attributes(SDE)  # todo: add more variables and corrseponding initializatons

    def printName(self):
        """ prints ItemID and item name"""
        print(self.itemName + ' has ID: ' + str(self.itemID))

    def initialize_attributes(self, sde):
        """  """
        self.itemName = sde.invTypes[self.itemID]['typeName']


class Blueprint(EVEItem):  # todo: fix import of SDE in EVEitem and Blueprint
    """ a blueprint class """

    def __init__(self, blueprintID, SDE):
        """ """
        super(Blueprint, self).__init__(blueprintID, SDE)
        self.bpID = blueprintID
        self.activities = {'copying', 'manufacturing', 'research_material', 'research_time'}
        self.blueprintTypeID = blueprintID
        self.maxProductionLimit = 0
        self.product = 0
        self.product_quantity = 0

    def fetch_bp_data(self):
        """ load info from SDE """
        # if sde == None:
        #     sde = SDE()
        #     print('Loading Blueprints to SDE')
        #     sde.import_pickle('blueprints')
        try:
            bpDict = SDE.blueprints[self.bpID]
        except KeyError:
            print('invalid BP ID')

        for key in bpDict:
            setattr(self, key, bpDict[key])

        self.product = self.activities['manufacturing']['products'][0]['typeID']
        self.product_quantity = self.activities['manufacturing']['products'][0]['quantity']

    def get_manufacturing_materials(self, runs=1, out='ID'):
        """ returns materials required for n runs"""
        requirement_list = self.activities['manufacturing']['materials']
        requirement_dict = {}
        for item in requirement_list:
            quantity = item['quantity'] * runs
            if out == 'ID':
                requirement_dict[item['typeID']] = [quantity]
            elif out == 'name':
                name = SDE.invTypes[item['typeID']]['typeName']
                requirement_dict[name] = quantity
        return requirement_dict




# %% Database Management

class SDE(object):
    """ Contains a series of dictionaries containing necessary sde data"""

    def __init__(self):

        self.DB_LOCATION_CSV = 'D:/Documents/py_code/EVE_database/csv/'
        self.DB_LOCATION_YAML = 'D:/Documents/py_code/EVE_database/sde/fsd/'
        self.DB_LOCATION_PICKLE = 'D:/Documents/py_code/EVE_database/'
        self.QUICK_IMPORT_LIST = ['invTypes','invCategories','invGroups','invMetaTypes','invMetaGroups','blueprints']

        self.db_list = []

        self.blueprints = None
        self.invCategories = None
        self.invGroups = None
        self.invMetaGroups = None
        self.invMetaTypes = None
        self.invTypes = None


        self.db_coreIndy = ['invTypes',
                            'invCategories',
                            'invGroups',
                            'invMetaTypes',
                            'invMetaGroups',
                            'blueprints']

        self.db_coreIndy_old = ['invTypes',
                                'invCategories',
                                'invGroups',
                                'invMetaTypes',
                                'invMetaGroups',
                                'industryActivity',
                                'industryActivityMaterials',
                                'industryActivityProbabilities']

        self.activityTypeID = {0: 'None',
                               1: 'Manufacturing',
                               2: 'Researching Technology',
                               3: 'Researching Time Efficiency',
                               4: 'Researching Material Efficiency',
                               5: 'Copying',
                               6: 'Duplicating',
                               7: 'Reverse Engineering',
                               8: 'Invention'}

    def import_data(self, dbName):
        """ imports data from a database file (csv or yaml)"""
        timer = gfs.Timer()
        timer.tic()
        if dbName in ['blueprints', 'typeIDs', 'groupIDs', 'graphicIDs']:
            print('importing: ' + dbName +  '.yaml')
            data = self.get_data_yaml(dbName)
        else:
            print('importing: ' + dbName + '.csv')
            data = self.get_data_csv(dbName)
        setattr(self, dbName, data)
        dt = timer.toc(out='return')
        print('Imported {0} in {1:.3f} ms'.format(dbName,dt))

    def export_pickle(self, dbName):
        """ create a pickle for each database loaded"""
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        try:
            timer = gfs.Timer()
            timer.tic()
            data = getattr(self, dbName)
            with open(filePath, 'wb+') as f:
                pickle.dump(data, f)
            dt = timer.toc(out='return')
            print('Exported {0} in {1:.3f} ms'.format(dbName, dt))
        except AttributeError:
            print('no data found with name ' + dbName)

    def export_multiple_pickle(self,dbList):
        """ import from sde and export to pickle the list of databases given """
        for db_name in dbList:
            print('importing: ' + db_name)
            self.import_data(db_name)
            print('exporting: ' + db_name)
            self.export_pickle(db_name)

    def import_pickle(self, dbName):  # todo: untested
        """ imports data from pickle dumped file"""
        # try:
        timer = gfs.Timer()
        timer.tic()
        print('Importing: ' + dbName)
        filePath = self.DB_LOCATION_PICKLE + dbName + '.p'
        with open(filePath, 'rb') as f:
            data = pickle.load(f)
            setattr(self, dbName, data)
        dt = timer.toc(out='return')
        print('Imported {0} in {1:.3f} ms'.format(dbName,dt))

    def import_indy_db(self):
        """ import databases relevant for indy calculator"""
        for dbName in self.QUICK_IMPORT_LIST:
            self.import_pickle(dbName)

    def get_filepath(self, db_name, ext):
        """
        :return : returns str of database file path.
        :param db_name: name of database file to load
        :type ext: str representing db_name file format

        """
        if ext == 'yaml':
            return self.DB_LOCATION_CSV + db_name + '.' + ext
        elif ext == 'csv':
            return self.DB_LOCATION_YAML + db_name + '.' + ext
        else:
            return 'Incorrect file extension.'

    def get_data_yaml(self, db_name):
        """ imports a YAML database to same name attribute"""
        with  open(self.DB_LOCATION_YAML + db_name + '.yaml', 'r') as f:
            data = yaml.load(f)
        return data

    def get_data_csv(self, db_name):
        """ imports a CSV database to same name attribute """
        headline = True
        data = {}
        with open(self.DB_LOCATION_CSV + db_name + '.csv', 'r', encoding="utf8") as f:
            reader = csv.reader(f)

            for line in reader:  # get headers for dict key names
                if headline:
                    headers = line
                    headline = False
                else:
                    key = int(line[0])
                    pos = 0

                    try:  # try appending new values to repeated 1st column items
                        # otherwise just add new entry
                        _ = data[key]
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = getNum_or_Str(word)
                            try:
                                data[key][label].append(val)
                            except AttributeError:
                                tmp = data[key][label]
                                data[key][label] = [tmp]
                                data[key][label].append(val)

                    except KeyError:
                        data[key] = {}
                        for word in line:
                            label = headers[pos]
                            pos += 1
                            val = gfs.getNum_or_Str(word)

                            data[key][label] = val
        return (data)

    # --------------- Old and deprectated methods ------------------

    def init_db_list(self):
        """ list all csv database files"""
        file_list = os.listdir(self.DB_LOCATION_CSV)
        for name in file_list:
            self.db_list.append(name[:-4])

    def import_all_csv(self):
        """ imports all csv database files"""
        if len(self.db_list) == 0:
            self.init_db_list()

        for name in self.db_list:
            print(name)
            self.import_data(self.get_filepath(name))

    def import_db_core_indy(self):
        """ import database files used for industry purposes."""
        for name in self.db_coreIndy_old:
            self.import_data(name)

    def open_file_xlsx(self, file):
        ''' '''
        wb = load_workbook(file)
        ws = wb.active
        self.name = ws['A1'].value
        # get column headers
        for col in ws.iter_cols(max_row=1):
            for cell in col:
                self.headers.append(cell.value)

        # write to data
        row_i = 0

        for row in ws.iter_rows(min_row=2):

            row_i += 1
            key = ws.cell(row=row_i, column=1).value
            col_i = 0
            self.data[key] = {}
            for header in self.headers:
                col_i += 1
                self.data[key][header] = ws.cell(row=row_i, column=col_i).value


if __name__ == '__main__':
    main()
